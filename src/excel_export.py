import os

import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment


def auto_adjust_columns(sheet):

    for column_cells in sheet.columns:

        length = 0

        column = column_cells[0].column_letter

        for cell in column_cells:

            try:

                if len(str(cell.value)) > length:

                    length = len(str(cell.value))

            except:

                pass

        adjusted_width = length + 5

        sheet.column_dimensions[column].width = adjusted_width


def style_sheet(sheet):

    header_fill = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    for cell in sheet[1]:

        cell.fill = header_fill

        cell.font = header_font

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    sheet.freeze_panes = "A2"

    auto_adjust_columns(sheet)


def export_to_excel(
    records,
    output_path="outputs/crash_reports.xlsx"
):

    try:

        os.makedirs(
            os.path.dirname(output_path),
            exist_ok=True
        )

        dataset_df = pd.DataFrame(records)

        first_record = records[0]

        report_df = pd.DataFrame({
            "Field": list(first_record.keys()),
            "Value": list(first_record.values())
        })

        with pd.ExcelWriter(
            output_path,
            engine="openpyxl"
        ) as writer:

            dataset_df.to_excel(
                writer,
                sheet_name="structured_dataset",
                index=False
            )

            report_df.to_excel(
                writer,
                sheet_name="formatted_report",
                index=False
            )

        workbook = load_workbook(output_path)

        dataset_sheet = workbook[
            "structured_dataset"
        ]

        report_sheet = workbook[
            "formatted_report"
        ]

        style_sheet(dataset_sheet)

        style_sheet(report_sheet)

        workbook.save(output_path)

        print(
            f"\n[INFO] Excel exported to: {output_path}"
        )

    except Exception as e:

        print(
            f"\n[ERROR] Excel export failed: {e}"
        )